import boto3
import argparse
import openpyxl
import os
from botocore.exceptions import ClientError
from sys import exit


class EncryptEC2:

    def __init__(self, instance_id: str, region: str, profile: str = 'default', key: str = 'alias/aws/ebs'):
        session = boto3.session.Session(profile_name=profile, region_name=region)

        self.instance_id = instance_id
        self.key = key
        self._ec2_client = session.client('ec2')
        self._ec2_details = self._ec2_client.describe_instances(InstanceIds=[instance_id])
        self._ec2_resource = session.resource('ec2')
        self._ec2_stop_waiter = self._ec2_client.get_waiter('instance_stopped')
        self._ebs_available_waiter = self._ec2_client.get_waiter('volume_available')
        self._snapshot_created_waiter = self._ec2_client.get_waiter('snapshot_completed')
        self._ec2_status_check_waiter = self._ec2_client.get_waiter('instance_status_ok')

        self._delay = 5
        self._max_attempts = 60

        if self.pre_checks():
            self._describe_ec2 = self._ec2_client.describe_instances(InstanceIds=[instance_id])
            for reservation in self._describe_ec2['Reservations']:
                self._instance_details = reservation['Instances'][0]
        else:
            # Exits the whole execution if pre-checks fails
            exit()

    def get_ebs_list(self) -> list[dict]:
        """Returns list of unencrypted volume details"""
        volume_ids = []
        for ebs in self._ec2_details['Reservations'][0]['Instances'][0]['BlockDeviceMappings']:
            volume_id = ebs['Ebs']['VolumeId']
            resp = self._ec2_client.describe_volumes(
                VolumeIds=[volume_id]
            )
            for volume in resp['Volumes']:
                if volume['Encrypted']:
                    pass
                else:
                    volume_ids.append(resp['Volumes'][0])
        return volume_ids

    def pre_checks(self) -> bool:
        """ Checks if the EC2 exists or supports encrypted EBS volumes. Need to add check if all EBS are encrypted"""
        try:
            # Checks if all volumes are already encrypted.
            if self.get_ebs_list():
                # Function : get_ebs_list returns list of unencrypted volumes. Thus if it returns, this check passed.
                pass
            else:
                # get_ebs_list returns nothing and thus meaning there are no unencrypted EBS volumes present
                print("All volumes are already encrypted")
                return False

            # Checks if the instance type supports EBS encryption
            instance_type = self._ec2_details['Reservations'][0]['InstanceType']
            if instance_type.startswith(('c1', 'm1', 'm2', 't1')):
                print(f"Instance type {instance_type} is not supported for encryption")
                return False
            else:
                print("Pre checks passed")
                return True

        except ClientError as err:
            if err.response['Error']['Code'] == 'InvalidInstanceID.Malformed':
                raise Exception("Instance not found")

    def get_az(self) -> str:
        """ Returns availability zone of the instance """
        return self._instance_details['Placement']['AvailabilityZone']

    def stop_instance(self) -> None:
        """ Stops EC2 Instance and wait for it to be in stopped state"""
        print(f"Stopping {self.instance_id}")
        self._ec2_client.stop_instances(
            InstanceIds=[
                self.instance_id,
            ]
        )
        self._ec2_stop_waiter.wait(
            InstanceIds=[
                self.instance_id,
            ],
            WaiterConfig={
                'Delay': self._delay,
                'MaxAttempts': self._max_attempts
            }
        )
        print(f"{self.instance_id} stopped")

    def detach_volume(self):
        """ Detach the EBS volumes attached to EC2, if they are unencrypted"""
        volumes = self.get_ebs_list()
        for item in volumes:
            volume_id = item['VolumeId']
            volume_type = item['VolumeType']
            self._ec2_client.create_tags(
                Resources=[
                    volume_id,
                ],
                Tags=[
                    {
                        'Key': 'device-name',
                        'Value': item['DeviceName']
                    },
                    {
                        'Key': 'volume-type',
                        'Value': volume_type
                    }
                ]
            )
            print(f"Detaching {volume_id} : {item['DeviceName']}")
            self._ec2_client.detach_volume(
                InstanceId=self.instance_id,
                VolumeId=volume_id
            )
            self._ebs_available_waiter.wait(
                VolumeIds=[
                    volume_id,
                ],
                WaiterConfig={
                    'Delay': self._delay,
                    'MaxAttempts': self._max_attempts
                }
            )

    def create_snapshots(self, volume_ids: list) -> list:
        created_snapshots = []
        for volume in volume_ids:
            volume_resource = self._ec2_resource.Volume(volume)

            # Pulling tags and excluding the ones that starts with keyword "aws" as aws blocks creating these manually
            final_tags = []
            for tag in volume_resource.tags:
                if tag['Key'].startswith('aws'):
                    pass
                else:
                    final_tags.append(tag)
            response = self._ec2_client.create_snapshot(
                VolumeId=volume,
                TagSpecifications=[
                    {
                        'ResourceType': 'snapshot',
                        'Tags': final_tags
                    }
                ]
            )
            self._snapshot_created_waiter.wait(
                SnapshotIds=[
                    response['SnapshotId'],
                ],
                WaiterConfig={
                    'Delay': self._delay,
                    'MaxAttempts': self._max_attempts
                }
            )
            created_snapshots.append(response['SnapshotId'])
            print(f"{response['SnapshotId']} created")
        return created_snapshots

    def create_volume(self, snapshot_ids: list, availability_zone: str) -> list:
        created_volumes = []
        for snapshot in snapshot_ids:
            snapshot_resource = self._ec2_resource.Snapshot(snapshot)
            volume_type = ''
            for tag in snapshot_resource.tags:
                if tag['Key'] == 'volume-type':
                    volume_type = tag['Value']

            # Due to SCP restrictions, we can't create gp2 or io1 or standard type EBS volumes, thus converting
            if volume_type == 'gp2' or volume_type == 'io1' or volume_type == 'standard':
                volume_type = 'gp3'

            response = self._ec2_client.create_volume(
                AvailabilityZone=availability_zone,
                Encrypted=True,
                KmsKeyId=self.key,
                SnapshotId=snapshot,
                VolumeType=volume_type,
                TagSpecifications=[
                    {
                        'ResourceType': 'volume',
                        'Tags': snapshot_resource.tags
                    }
                ]
            )
            self._ebs_available_waiter.wait(
                VolumeIds=[
                    response['VolumeId'],
                ],
                WaiterConfig={
                    'Delay': self._delay,
                    'MaxAttempts': self._max_attempts
                }
            )
            created_volumes.append(response['VolumeId'])
        print(f"{created_volumes} created")
        return created_volumes

    def attach_volume(self, volume_ids: list) -> bool:
        for volume in volume_ids:
            volume_resource = self._ec2_resource.Volume(volume)
            device_name = ''
            for tag in volume_resource.tags:
                if tag['Key'] == 'device-name':
                    device_name = tag['Value']
            self._ec2_client.attach_volume(
                Device=device_name,
                InstanceId=self.instance_id,
                VolumeId=volume
            )
        print(f"{volume_ids} attached")
        return True

    def start_instance(self):
        self._ec2_client.start_instances(InstanceIds=[self.instance_id])
        print(f"{self.instance_id} started")
        self._ec2_status_check_waiter.wait(
            InstanceIds=[
                self.instance_id,
            ],
            WaiterConfig={
                'Delay': self._delay,
                'MaxAttempts': self._max_attempts
            }
        )
        print(f"{self.instance_id} has passed 2/2 status checks")

    def delete_snapshots(self, snapshot_list: list):
        for item in snapshot_list:
            self._ec2_client.delete_snapshot(
                SnapshotId=item
            )
            print(item, "has been deleted")

    def start_encryption(self):
        availability_zone = self.get_az()
        self.stop_instance()
        self.detach_volume()
        volume_ids = [ebs['VolumeId'] for ebs in self.get_ebs_list()]
        snapshots = self.create_snapshots(volume_ids=volume_ids)
        encrypted_volumes = self.create_volume(snapshot_ids=snapshots, availability_zone=availability_zone)
        self.attach_volume(volume_ids=encrypted_volumes)
        self.start_instance()
        self.delete_snapshots(snapshot_list=snapshots)
        self.start_instance()


def parse_arguments():
    parser = argparse.ArgumentParser(description='Encrypt EC2 instance')
    sub_parser = parser.add_subparsers(dest='sub_command')

    bulk = sub_parser.add_parser('bulk', help="For bulk execution Excel file path with instance details")
    single = sub_parser.add_parser('single', help="For single instance ")

    bulk.add_argument('-f', '--file', metavar='input-file.xlsx', nargs='?', required=True)
    single.add_argument('-i', '--instance', help="Single instance id", required=True)
    single.add_argument('-r', '--region', help="AWS region", required=True)
    single.add_argument('-p', '--profile', help="AWS profile", default='default')
    single.add_argument('-k', '--key', help="KMS key id [optional]. If not provided, AWS managed key will be used")
    return parser.parse_args()


def file_exists(file):
    if os.path.exists(file):
        return True
    else:
        raise Exception("Oops, that path doesn't exist")


def bulk_execution(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    heading = []
    first_row = ws.iter_rows(max_row=1, values_only=True)
    for i in first_row:
        heading.extend(list(i))
    instance_row = heading.index('InstanceID')
    region_row = heading.index('Region')
    account_row = heading.index('Account')
    try:
        key_row = heading.index('Key')
    except ValueError:
        key_row = None
    data = ws.iter_rows(min_row=2, values_only=True)
    for row in data:
        instance_id = row[instance_row]
        region = row[region_row]
        account = row[account_row]
        print(instance_id, region, account)
        if key_row:
            key = row[key_row]
            EncryptEC2(instance_id=instance_id, region=region, profile=account, key=key).start_encryption()
        else:
            print("Proceeding without custom key, AWS managed key will be used for encryption")
            EncryptEC2(instance_id=instance_id, region=region, profile=account).start_encryption()


def single_execution(args):
    instance_id = args.instance
    region = args.region
    account = args.profile
    try:
        key = args.key
    except ValueError:
        key = None
    if key:
        EncryptEC2(instance_id=instance_id, region=region, profile=account, key=key).start_encryption()
    else:
        EncryptEC2(instance_id=instance_id, region=region, profile=account).start_encryption()


if __name__ == "__main__":
    arguments = parse_arguments()
    if arguments.sub_command == 'bulk':
        if file_exists(arguments.file):
            bulk_execution(arguments.file)
        else:
            raise Exception("Oops, that path doesn't exist")
    if arguments.sub_command == 'single':
        single_execution(arguments)
