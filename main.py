from InquirerPy import inquirer
from InquirerPy.base.control import Choice
from InquirerPy.validator import PathValidator, EmptyInputValidator
from InquirerPy import get_style
import openpyxl
import os
from ec2 import EncryptEC2
from rds import EncryptRDS
from efs import EncryptEFS
import time
from boto3.session import Session


def bulk_execution(file, resource_type):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    heading = []
    first_row = ws.iter_rows(max_row=1, values_only=True)
    for i in first_row:
        heading.extend(list(i))
    resource_row = heading.index('resourceid')
    region_row = heading.index('region')
    profile_row = heading.index('profile')
    try:
        key_row = heading.index('Key')
    except ValueError:
        key_row = None
    data = ws.iter_rows(min_row=2, values_only=True)
    for row in data:
        resource_id = row[resource_row]
        region = row[region_row]
        profile = row[profile_row]
        print(resource_id, region, profile)

        if resource_type == 'EC2':
            if key_row:
                key = row[key_row]
                EncryptEC2(instance_id=resource_id, region=region, profile=profile, key=key).start_encryption()
            else:
                print("Proceeding without custom key, AWS managed key will be used for encryption")
                EncryptEC2(instance_id=resource_id, region=region, profile=profile).start_encryption()

        elif resource_type == 'EBS':
            if key_row:
                key = row[key_row]
                EncryptEC2(instance_id=resource_id, region=region, profile=profile, key=key).start_encryption()
            else:
                print("Proceeding without custom key, AWS managed key will be used for encryption")
                EncryptEC2(instance_id=resource_id, region=region, profile=profile).start_encryption()

        elif resource_type == 'RDS':
            if key_row:
                key = row[key_row]
                EncryptRDS(rds_identifier=resource_id, region=region, profile=profile, key=key).start_encryption()
            else:
                print("Proceeding without custom key, AWS managed key will be used for encryption")
                EncryptRDS(rds_identifier=resource_id, region=region, profile=profile).start_encryption()

        elif resource_type == 'EFS':
            if key_row:
                key = row[key_row]
                EncryptEFS(efs_id=resource_id, region=region, profile=profile, key=key).start_encryption()
            else:
                print("Proceeding without custom key, AWS managed key will be used for encryption")
                EncryptEFS(efs_id=resource_id, region=region, profile=profile).start_encryption()

        else:
            pass


def single_execution(resource_type, resource_id, region, profile, key=None):

    if resource_type == 'EC2':
        if key:
            EncryptEC2(instance_id=resource_id, region=region, profile=profile, key=key).start_encryption()
        else:
            print("Proceeding without custom key, AWS managed key will be used for encryption")
            EncryptEC2(instance_id=resource_id, region=region, profile=profile).start_encryption()

    elif resource_type == 'EBS':
        if key:
            EncryptEC2(instance_id=resource_id, region=region, profile=profile, key=key).start_encryption()
        else:
            print("Proceeding without custom key, AWS managed key will be used for encryption")
            EncryptEC2(instance_id=resource_id, region=region, profile=profile).start_encryption()

    elif resource_type == 'RDS':
        if key:
            EncryptRDS(rds_identifier=resource_id, region=region, profile=profile, key=key).start_encryption()
        else:
            print("Proceeding without custom key, AWS managed key will be used for encryption")
            EncryptRDS(rds_identifier=resource_id, region=region, profile=profile).start_encryption()

    elif resource_type == 'EFS':
        if key:
            EncryptEFS(efs_id=resource_id, region=region, profile=profile, key=key).start_encryption()
        else:
            print("Proceeding without custom key, AWS managed key will be used for encryption")
            EncryptEFS(efs_id=resource_id, region=region, profile=profile).start_encryption()
    else:
        pass


def resource_id_text_parser(resource_type):
    resource_text = 'Resource ID'
    if resource_type == 'EC2':
        resource_text = 'Instance ID'
    elif resource_type == 'EBS':
        resource_text = 'Volume ID'
    elif resource_type == 'RDS':
        resource_text = 'RDS Identifier'
    elif resource_type == 'EFS':
        resource_text = 'Filesystem ID'
    return resource_text


def aws_profile_completer():
    profiles = Session().available_profiles
    profiles_dict = {profile: None for profile in profiles}
    return profiles_dict


def aws_region_completer(resource_type: str):
    service_name = resource_type.lower()
    regions = Session().get_available_regions(service_name)
    regions_dict = {region: None for region in regions}
    return regions_dict


def main():
    style = get_style({"questionmark": "fg:yellow", "answer": "#000000", "pointer": "orange bold", "question": "green bold"}, style_override=False)
    proceed = True
    while proceed:
        try:
            resource_type = inquirer.rawlist(
                style=style,
                message="Select the resource type to encrypt:",
                #  Choices below must be exact name of the AWS service for downstream dependencies
                choices=[
                    "EC2",
                    "EBS",
                    "RDS",
                    "EFS",
                    Choice(value=None, name="Exit"),
                ],
                default=1,
            ).execute()

            if resource_type:
                count = inquirer.rawlist(
                    message="Select the resource type to encrypt:",
                    choices=[
                        "Single",
                        "Bulk",
                        Choice(value=None, name="Exit"),
                    ],
                    default=1
                ).execute()

                if count == 'Single':
                    resource_text = resource_id_text_parser(resource_type)
                    resource_id = inquirer.text(
                        message=f"Enter the {resource_text}:",
                        validate=EmptyInputValidator("Input should not be empty")
                    ).execute()
                    region = inquirer.text(
                        message="Enter the AWS region code:",
                        completer=aws_region_completer(resource_type),
                        validate=EmptyInputValidator("Input should not be empty"),
                        multicolumn_complete=True,
                    ).execute()

                    profile = inquirer.text(
                        message="Enter the AWS profile (Leave blank for default):",
                        filter=lambda result: 'default' if result == '' else result,  # makes the profile default if left blank
                        completer=aws_profile_completer(),
                        mandatory=False
                        # qmark='?'  -- to change the qmark at beginning
                    ).execute()

                    key = inquirer.text(message="Enter the custom KMS key (Leave blank to use AWS managed key):").execute()
                    if key:
                        single_execution(resource_type=resource_type, resource_id=resource_id, region=region, profile=profile,
                                         key=key)
                    else:
                        single_execution(resource_type=resource_type, resource_id=resource_id, region=region, profile=profile)

                elif count == 'Bulk':
                    home_path = "~/" if os.name == "posix" else "C:\\"
                    file_path = inquirer.filepath(
                        message="Enter the excel file path:",
                        default=home_path,
                        validate=lambda result: PathValidator(is_file=True, message="Input is not a file") and result[-5:] in ['.xlsx', 'xlsm', 'xltx', 'xltm'],
                        invalid_message="Input is not a file or with wrong extension. Please select an excel (.xlsx) file."
                    ).execute()
                    bulk_execution(file=file_path, resource_type=resource_type)

                else:
                    proceed = False
                    continue

                proceed = inquirer.confirm(message="Do you want to proceed again?", default=True).execute()
                time.sleep(1)

            else:
                proceed = False

        except KeyboardInterrupt:
            print("Exiting")
            time.sleep(1)
            proceed = False


if __name__ == "__main__":
    main()
